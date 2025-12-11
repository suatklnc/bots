#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import logging
import json
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton
from telegram.constants import ParseMode
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    ConversationHandler,
    filters,
)

# --- Ayarlar ---
TOKEN = "8437498731:AAFU6_v8gc-yi4AuMJLsZs_gIJ7y2D7oYK8"
STATE_FILE = "bot_state.json"
EXCEL_FOLDER = "excels"

# --- Logging ---
logging.basicConfig(format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

# --- Conversation states ---
ASK_NAME, ASK_AUTHOR, ASK_PLACE, ASK_PUBLISHER, ASK_DATE = range(5)

# --- Yardımcı fonksiyonlar ---
def ensure_dirs():
    if not os.path.exists(EXCEL_FOLDER):
        os.makedirs(EXCEL_FOLDER)

def load_state():
    if not os.path.exists(STATE_FILE):
        return {}
    with open(STATE_FILE, "r", encoding="utf-8") as f:
        try:
            return json.load(f)
        except Exception:
            return {}

def save_state(state):
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)

def auto_resize(ws):
    # sütun genişliklerini en uzun hücreye göre ayarlar
    for col in ws.columns:
        max_len = 0
        try:
            col_letter = col[0].column_letter
        except Exception:
            continue
        for cell in col:
            if cell.value is not None:
                l = len(str(cell.value))
                if l > max_len:
                    max_len = l
        ws.column_dimensions[col_letter].width = max_len + 2

def create_new_excel(chat_id):
    ensure_dirs()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = os.path.join(EXCEL_FOLDER, f"kitaplar_{chat_id}_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "No"
    ws["B1"] = "Kitap Adı"
    ws["C1"] = "Yazar"
    ws["D1"] = "Yayın Yeri"
    ws["E1"] = "Yayın Evi"
    ws["F1"] = "Yayın Tarihi"

    auto_resize(ws)
    wb.save(filename)

    state = load_state()
    state[str(chat_id)] = {"filename": filename, "last_row": 1}
    save_state(state)
    return filename

def get_chat_state(chat_id):
    return load_state().get(str(chat_id))

def set_chat_state(chat_id, filename, last_row):
    state = load_state()
    state[str(chat_id)] = {"filename": filename, "last_row": last_row}
    save_state(state)

def append_row(chat_id, name, author, place, pub, date):
    s = get_chat_state(chat_id)
    if not s:
        filename = create_new_excel(chat_id)
        last_row = 1
    else:
        filename = s["filename"]
        last_row = s.get("last_row", 1)

    wb = load_workbook(filename)
    ws = wb.active

    next_row = last_row + 1
    ws.cell(row=next_row, column=1, value=next_row - 1)
    ws.cell(row=next_row, column=2, value=name)
    ws.cell(row=next_row, column=3, value=author)
    ws.cell(row=next_row, column=4, value=place)
    ws.cell(row=next_row, column=5, value=pub)
    ws.cell(row=next_row, column=6, value=date)

    auto_resize(ws)
    wb.save(filename)

    set_chat_state(chat_id, filename, next_row)
    return next_row - 1

def delete_last(chat_id):
    s = get_chat_state(chat_id)
    if not s:
        return False, "Herhangi bir Excel dosyası bulunamadı."
    filename = s["filename"]
    if not os.path.exists(filename):
        return False, "Excel dosyası bulunamadı."
    wb = load_workbook(filename)
    ws = wb.active
    last_row = s.get("last_row", 1)
    if last_row <= 1:
        return False, "Silinecek kayıt yok."
    ws.delete_rows(last_row)
    wb.save(filename)
    set_chat_state(chat_id, filename, last_row - 1)
    return True, "Son kayıt başarıyla silindi."

def get_books_list(chat_id):
    """Excel dosyasından kitapları okuyup liste döndürür"""
    s = get_chat_state(chat_id)
    if not s:
        return None, "Herhangi bir Excel dosyası bulunamadı."
    filename = s["filename"]
    if not os.path.exists(filename):
        return None, "Excel dosyası bulunamadı."
    
    try:
        wb = load_workbook(filename)
        ws = wb.active
        books = []
        # 2. satırdan başla (1. satır başlık)
        for row in range(2, ws.max_row + 1):
            no = ws.cell(row=row, column=1).value
            name = ws.cell(row=row, column=2).value
            author = ws.cell(row=row, column=3).value
            place = ws.cell(row=row, column=4).value
            pub = ws.cell(row=row, column=5).value
            date = ws.cell(row=row, column=6).value
            
            if name:  # Kitap adı varsa listeye ekle
                books.append({
                    "no": no,
                    "name": name or "",
                    "author": author or "",
                    "place": place or "",
                    "publisher": pub or "",
                    "date": date or ""
                })
        return books, None
    except Exception as e:
        return None, f"Hata: {str(e)}"

# --- Komut / Mesaj akışları ---
WELCOME_TEXT = (
    "Esselamu aleyküm, hoş geldiniz 🌿\n\n"
    "Bu bot, İŞKUR kapsamında Selçuk İlahiyat için kitap kayıt işlemlerinizi kolaylaştırmak üzere hazırlanmıştır.\n\n"
    "📚 *Öne çıkan özellikler (hepsi aktif):*\n"
    "- \"Yeni Dosya\" butonuna basarak kitap eklemeye başlayabilirsiniz.\n"
    "- Kayıtlar Excel dosyasına otomatik kaydedilir (A: No, B: Kitap Adı, C: Yazar, D: Yayın Yeri, E: Yayın Evi, F: Yayın Tarihi).\n"
    "- \"Bitti\" butonu ile kitap ekleme işlemini sonlandırabilirsiniz.\n"
    "- \"Kitapları Listele\" ile kayıtlı kitapları görüntüleyebilirsiniz.\n"
    "- \"Excel'i İndir\" ile Excel dosyanızı Telegram üzerinden indirebilirsiniz.\n"
    "- \"Son Exceli Sil\" komutuyla en son kaydı silebilirsiniz.\n"
    "✨ İşlemlere başlamak için \"Yeni Dosya\" butonuna basın.\n"
)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    # varsa Excel hazırla yoksa oluştur (kullanıcı daha önce dosya oluşturmadıysa)
    if not get_chat_state(chat_id):
        create_new_excel(chat_id)
    # gönderilecek klavye: kullanıcı komutlarını hızlıca yazabilsin
    keyboard = ReplyKeyboardMarkup([
        ["Yeni Dosya"],
        ["Kitapları Listele", "Excel'i İndir"],
        ["Son Exceli Sil"]
    ], resize_keyboard=True)
    await update.message.reply_text(WELCOME_TEXT, reply_markup=keyboard, parse_mode=ParseMode.MARKDOWN)
    # Döngüye sokmuyoruz, sadece hoş geldin mesajı
    return ConversationHandler.END

# Yeni Dosya mesajı ile de akışı başlat/yeniden başlat
async def new_file_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    create_new_excel(chat_id)
    # Bitti butonu göstermiyoruz, sadece ilk soru
    await update.message.reply_text("📘 Yeni dosya oluşturuldu. Yeni kayda başlıyorum.\n1) Kitap adı nedir?")
    return ASK_NAME

# Sorular akışı
async def ask_author(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["name"] = update.message.text.strip()
    # Bitti butonu göstermiyoruz
    await update.message.reply_text("2) Yazar?")
    return ASK_AUTHOR

async def ask_place(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["author"] = update.message.text.strip()
    # Bitti butonu göstermiyoruz
    await update.message.reply_text("3) Yayın yeri?")
    return ASK_PLACE

async def ask_publisher(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["place"] = update.message.text.strip()
    # Bitti butonu göstermiyoruz
    await update.message.reply_text("4) Yayın evi?")
    return ASK_PUBLISHER

async def ask_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["publisher"] = update.message.text.strip()
    # Son soru - Bitti butonu yok, sadece soru
    await update.message.reply_text("5) Yayın tarihi?")
    return ASK_DATE

async def save_and_continue(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["date"] = update.message.text.strip()
    chat_id = update.effective_chat.id

    # kaydet
    no = append_row(
        chat_id,
        context.user_data.get("name", ""),
        context.user_data.get("author", ""),
        context.user_data.get("place", ""),
        context.user_data.get("publisher", ""),
        context.user_data.get("date", "")
    )

    # temizle geçiş için
    context.user_data.clear()

    # Başarı mesajı ve Bitti butonu
    keyboard = ReplyKeyboardMarkup([["Bitti"]], resize_keyboard=True)
    await update.message.reply_text(f"✅ {no}. kitap kaydedildi!\n\nYeni kitap eklemek için bilgileri girin veya \"Bitti\" butonuna basın.\n1) Kitap adı nedir?", reply_markup=keyboard)
    return ASK_NAME

# "Bitti" handler - döngüden çıkış ve Excel gönderme
async def finish_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    context.user_data.clear()
    
    # Excel dosyasını otomatik olarak gönder
    s = get_chat_state(chat_id)
    if s and os.path.exists(s.get("filename", "")):
        filename = s["filename"]
        try:
            with open(filename, 'rb') as f:
                await update.message.reply_document(
                    document=f,
                    filename=os.path.basename(filename),
                    caption="📊 Excel dosyanız hazır!"
                )
        except Exception as e:
            logger.error(f"Excel gönderilirken hata: {str(e)}")
            await update.message.reply_text(f"❌ Dosya gönderilirken hata oluştu: {str(e)}")
    
    # Ana klavyeyi geri getir
    keyboard = ReplyKeyboardMarkup([
        ["Yeni Dosya"],
        ["Kitapları Listele", "Excel'i İndir"],
        ["Son Exceli Sil"]
    ], resize_keyboard=True)
    await update.message.reply_text("✅ Kitap ekleme işlemi tamamlandı. İsterseniz \"Yeni Dosya\" butonuna basarak tekrar başlayabilirsiniz.", reply_markup=keyboard)
    return ConversationHandler.END

# "Son Exceli Sil" handler
async def delete_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    ok, msg = delete_last(chat_id)
    # Ana klavyeyi geri getir
    keyboard = ReplyKeyboardMarkup([
        ["Yeni Dosya"],
        ["Kitapları Listele", "Excel'i İndir"],
        ["Son Exceli Sil"]
    ], resize_keyboard=True)
    await update.message.reply_text(msg, reply_markup=keyboard)
    # Döngüye girmiyoruz, ana menüye dönüyoruz
    return ConversationHandler.END

# "Son Bilgiyi Düzelt" -> basit yöntem: sil, ana menüye dön
async def fix_last_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    ok, msg = delete_last(chat_id)
    # Ana klavyeyi geri getir
    keyboard = ReplyKeyboardMarkup([
        ["Yeni Dosya"],
        ["Kitapları Listele", "Excel'i İndir"],
        ["Son Exceli Sil"]
    ], resize_keyboard=True)
    if not ok:
        await update.message.reply_text(msg, reply_markup=keyboard)
    else:
        await update.message.reply_text("✅ Son kayıt silindi. \"Yeni Dosya\" butonuna basarak yeniden başlayabilirsiniz.", reply_markup=keyboard)
    # Döngüye girmiyoruz, ana menüye dönüyoruz
    return ConversationHandler.END

# "Kitapları Listele" handler
async def list_books_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    books, error = get_books_list(chat_id)
    
    if error:
        await update.message.reply_text(error)
        return ASK_NAME
    
    if not books or len(books) == 0:
        await update.message.reply_text("📚 Henüz kayıtlı kitap bulunmuyor.")
        return ASK_NAME
    
    # Kitapları formatla
    message = "📚 *Kayıtlı Kitaplar:*\n\n"
    for book in books:
        message += f"*{book['no']}.* {book['name']}\n"
        message += f"   👤 Yazar: {book['author']}\n"
        message += f"   📍 Yer: {book['place']}\n"
        message += f"   🏢 Yayın Evi: {book['publisher']}\n"
        message += f"   📅 Tarih: {book['date']}\n\n"
    
    message += f"\n*Toplam {len(books)} kitap kayıtlı.*"
    
    # Telegram mesaj limiti 4096 karakter, eğer uzunsa böl
    if len(message) > 4000:
        # İlk mesajı gönder
        await update.message.reply_text(message[:4000] + "\n\n... (devam ediyor)", parse_mode=ParseMode.MARKDOWN)
        # Kalan kısmı gönder
        remaining = message[4000:]
        chunks = [remaining[i:i+4000] for i in range(0, len(remaining), 4000)]
        for chunk in chunks:
            await update.message.reply_text(chunk, parse_mode=ParseMode.MARKDOWN)
    else:
        await update.message.reply_text(message, parse_mode=ParseMode.MARKDOWN)
    
    return ASK_NAME

# "Excel'i İndir" handler
async def send_excel_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    s = get_chat_state(chat_id)
    
    if not s:
        await update.message.reply_text("❌ Herhangi bir Excel dosyası bulunamadı. Önce kitap ekleyin.")
        return ASK_NAME
    
    filename = s["filename"]
    if not os.path.exists(filename):
        await update.message.reply_text("❌ Excel dosyası bulunamadı.")
        return ASK_NAME
    
    try:
        with open(filename, 'rb') as f:
            await update.message.reply_document(
                document=f,
                filename=os.path.basename(filename),
                caption="📊 Excel dosyanız hazır!"
            )
    except Exception as e:
        await update.message.reply_text(f"❌ Dosya gönderilirken hata oluştu: {str(e)}")
    
    return ASK_NAME

# Fallback: eğer bir metin geldi ve bunlardan biri ise ilgili handler'ı çağır
async def general_text_fallback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip().lower()
    if text == "bitti":
        return await finish_handler(update, context)
    if text == "yeni dosya":
        return await new_file_handler(update, context)
    if text == "kitapları listele" or text == "liste":
        return await list_books_handler(update, context)
    if text == "excel'i indir" or text == "excel indir" or text == "dosyayı indir":
        return await send_excel_handler(update, context)
    if text == "son exceli sil" or text == "son excelı sil" or text == "son exceli sil.":
        return await delete_handler(update, context)
    # Eğer diğer metinlerse ConversationHandler sıradaki state'e yönlendirir (normal akış)
    return None

# --- Botu başlat ---
def main():
    ensure_dirs()
    app = ApplicationBuilder().token(TOKEN).build()

    conv = ConversationHandler(
        entry_points=[
            CommandHandler("start", start),
            MessageHandler(filters.Regex("(?i)^Yeni Dosya$"), new_file_handler),
            MessageHandler(filters.Regex("(?i)^Kitapları Listele$"), list_books_handler),
            MessageHandler(filters.Regex("(?i)^Excel'i İndir$"), send_excel_handler),
            MessageHandler(filters.Regex("(?i)^Son Exceli Sil$"), delete_handler),
        ],
        states={
            ASK_NAME: [
                MessageHandler(filters.Regex("(?i)^Bitti$"), finish_handler),
                MessageHandler(filters.Regex("(?i)^Yeni Dosya$"), new_file_handler),
                MessageHandler(filters.Regex("(?i)^Kitapları Listele$"), list_books_handler),
                MessageHandler(filters.Regex("(?i)^Excel'i İndir$"), send_excel_handler),
                MessageHandler(filters.TEXT & ~filters.COMMAND, ask_author),
            ],
            ASK_AUTHOR: [
                MessageHandler(filters.Regex("(?i)^Kitapları Listele$"), list_books_handler),
                MessageHandler(filters.Regex("(?i)^Excel'i İndir$"), send_excel_handler),
                MessageHandler(filters.TEXT & ~filters.COMMAND, ask_place)
            ],
            ASK_PLACE: [
                MessageHandler(filters.Regex("(?i)^Kitapları Listele$"), list_books_handler),
                MessageHandler(filters.Regex("(?i)^Excel'i İndir$"), send_excel_handler),
                MessageHandler(filters.TEXT & ~filters.COMMAND, ask_publisher)
            ],
            ASK_PUBLISHER: [
                MessageHandler(filters.Regex("(?i)^Kitapları Listele$"), list_books_handler),
                MessageHandler(filters.Regex("(?i)^Excel'i İndir$"), send_excel_handler),
                MessageHandler(filters.TEXT & ~filters.COMMAND, ask_date)
            ],
            ASK_DATE: [
                MessageHandler(filters.Regex("(?i)^Kitapları Listele$"), list_books_handler),
                MessageHandler(filters.Regex("(?i)^Excel'i İndir$"), send_excel_handler),
                MessageHandler(filters.TEXT & ~filters.COMMAND, save_and_continue)
            ],
        },
        fallbacks=[MessageHandler(filters.TEXT & ~filters.COMMAND, general_text_fallback)],
        allow_reentry=True,
    )

    app.add_handler(conv)

    # Standalone komutlar (ConversationHandler dışında da çalışsın)
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("liste", list_books_handler))
    app.add_handler(CommandHandler("dosya", send_excel_handler))

    logger.info("Bot başlatılıyor...")
    app.run_polling()

if __name__ == "__main__":
    main()
